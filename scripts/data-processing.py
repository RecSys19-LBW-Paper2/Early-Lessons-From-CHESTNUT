#!/usr/bin/python
# -*- coding: UTF-8 -*-

import openpyxl
import numpy as np


def read_file(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    names_1 = ws['A']     #list of name of the users
    rates_1 = ws['C']     #list of number of rated movies
    print('Row {}, Column {} is {}'.format(names_1.row, names_1.column, names_1.value))
    print('Cell {} is {}\n'.format(names_1.coordinate, names_1.value))
    print(len(names_1))

    wb = openpyxl.load_workbook('UserData.xlsx')
    ws = wb.active
    names_2 = ws['A']
    schools = ws['B']
    ages = ws['C']
    rates_2 = np.zeros(len(names_2))

    #two files are in different order, this step makes it ordered
    for i in range(len(names_1)):
        for j in range(len(names_2)):
            if names_1[i] == names_2[j]:
                rates_2[j] = rates_1[i]

    list_count = [0,0,0,0,0]
    list_rates_count = [0,0,0,0,0]
    list_ages_count = [0,0,0,0,0]
    for i in range(len(names_2)):
        if schools[i] == 'Campus 1':
            list_count[0] = list_count[0] + 1
            list_rates_count[0] = list_rates_count[0] + rates_2[i]
            list_ages_count[0] = list_ages_count[0] + 1
        elif schools[i] == 'Campus 2':
            list_count[1] = list_count[1] + 1
            list_rates_count[1] = list_rates_count[1] + rates_2[i]
            list_ages_count[1] = list_ages_count[1] + 1
        elif schools[i] == 'Campus 3':
            list_count[2] = list_count[2] + 1
            list_rates_count[2] = list_rates_count[2] + rates_2[i]
            list_ages_count[2] = list_ages_count[2] + 1
        elif schools[i] == 'Campus 4':
            list_count[3] = list_count[3] + 1
            list_rates_count[3] = list_rates_count[3] + rates_2[i]
            list_ages_count[3] = list_ages_count[3] + 1
        else:
            list_count[4] = list_count[4] + 1
            list_rates_count[4] = list_rates_count[4] + rates_2[i]
            list_ages_count[4] = list_ages_count[4] + 1

    all = list_count[0] + list_count[1] + list_count[2] + list_count[3] + list_count[4]
    list_avg_rates = [0,0,0,0,0]
    list_avg_ages = [0,0,0,0,0]
    for i in range(5):
        list_avg_rates[i] = float(list_rates_count[i]) / list_count[i]
        list_avg_ages[i] = float(list_ages_count[i]) / list_count[i]

    return list_avg_rates,list_avg_ages


def main():
    file_list = ['../results-from-CHESTNUT.xlsx','../results-from-ItemBased.xlsx','../results-from-UserBased.xlsx']
    for file in file_list:
        list_avg_rates,list_avg_ages = read_file(file)
        for i in range(5):
            print('avg_rates in %s from Campus %d: %f'%(file,i,list_avg_rates[i]))
            print('avg_ages in %s from Campus %d: %f'%(file,i,list_avg_ages[i]))